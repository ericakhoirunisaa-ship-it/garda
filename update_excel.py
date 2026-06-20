import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment, PatternFill
from PIL import Image
import os, io

FOTO_DIR = r"C:\xampp\htdocs\GardaSE\0. FOTO CALON MITRA STATISTIK TAMBAHAN"
XLSX_FILE     = r"C:\xampp\htdocs\GardaSE\data se.xlsx"
XLSX_OUT      = r"C:\xampp\htdocs\GardaSE\data se_updated.xlsx"

COL_STORAGE = 24
COL_SESUAI  = 25
COL_FOTO    = 26
IMG_W = 75
IMG_H = 100
ROW_H = 78  # points (~2.75 cm)

# Build case-insensitive photo dict: stem.lower() -> filename
photo_files = {}
for fname in os.listdir(FOTO_DIR):
    if fname.lower().endswith('.png'):
        stem = os.path.splitext(fname)[0]
        photo_files[stem.lower()] = fname

# Special override: nama di Excel → nama file foto
SPECIAL = {
    'MIFTAHUL JANNAH': 'MIFTAHUL JANNAH (2).png',
}

def find_photo(name):
    name = ' '.join(str(name).strip().split())  # normalize spaces
    if name in SPECIAL:
        return SPECIAL[name]
    return photo_files.get(name.lower())

def cek_sesuai(ram_val):
    if ram_val is None:
        return 'Tidak'
    s = str(ram_val).strip()
    if s in ('-', '', 'None'):
        return 'Tidak'
    try:
        ram = float(s)
        return 'Iya' if 3 <= ram <= 24 else 'Tidak'
    except:
        return 'Tidak'

def resize_img(path, w, h):
    with Image.open(path) as im:
        im.thumbnail((w, h), Image.LANCZOS)
        buf = io.BytesIO()
        im.save(buf, format='PNG')
        buf.seek(0)
        return buf

wb = openpyxl.load_workbook(XLSX_FILE)
ws = wb.active

# Headers
hdr_fill = PatternFill('solid', fgColor='F79039')
hdr_font = Font(bold=True, color='FFFFFF')
for col, label in [(COL_STORAGE, 'Storage HP (GB)'),
                   (COL_SESUAI,  'Apakah Sesuai (HP)'),
                   (COL_FOTO,    'Foto')]:
    c = ws.cell(1, col)
    c.value = label
    c.font  = hdr_font
    c.fill  = hdr_fill
    c.alignment = Alignment(horizontal='center', vertical='center')

ws.column_dimensions[openpyxl.utils.get_column_letter(COL_FOTO)].width = 12
ws.column_dimensions[openpyxl.utils.get_column_letter(COL_STORAGE)].width = 16
ws.column_dimensions[openpyxl.utils.get_column_letter(COL_SESUAI)].width = 20

ok = 0
no_photo = []

for row in range(2, ws.max_row + 1):
    nama = ws.cell(row, 1).value
    if not nama:
        continue

    # Apakah sesuai (berdasarkan RAM)
    ram = ws.cell(row, 23).value
    sesuai_cell = ws.cell(row, COL_SESUAI)
    sesuai_cell.value = cek_sesuai(ram)
    sesuai_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    if sesuai_cell.value == 'Iya':
        sesuai_cell.font = Font(color='16A34A', bold=True)
    else:
        sesuai_cell.font = Font(color='DC2626', bold=True)

    # Foto
    photo_fname = find_photo(nama)
    if photo_fname:
        img_path = os.path.join(FOTO_DIR, photo_fname)
        if os.path.exists(img_path):
            try:
                buf    = resize_img(img_path, IMG_W, IMG_H)
                xl_img = XLImage(buf)
                xl_img.width  = IMG_W
                xl_img.height = IMG_H
                col_letter = openpyxl.utils.get_column_letter(COL_FOTO)
                ws.add_image(xl_img, f'{col_letter}{row}')
                ws.row_dimensions[row].height = ROW_H
                ok += 1
            except Exception as e:
                no_photo.append(f'{nama}: {e}')
        else:
            no_photo.append(f'{nama}: file tidak ditemukan ({photo_fname})')
    else:
        no_photo.append(f'{nama}: tidak ada foto')

wb.save(XLSX_OUT)
print(f'Selesai. Foto berhasil: {ok}')
print(f'Tanpa foto ({len(no_photo)}):')
for n in no_photo:
    print(' -', n)
